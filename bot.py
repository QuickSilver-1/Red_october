from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart
from aiogram.types import Message, CallbackQuery, ReplyKeyboardRemove
from keyboards import *
from media import *
from config import config_1
from re import match
import openpyxl


dp = Dispatcher()
bot = Bot(token=config_1.TOKEN)

class Register(StatesGroup):
    fio = State()
    age = State()
    city = State()
    mail = State()
    number = State()

# @dp.message(F.document)
# async def photo_handler(message: Message) -> None:
#     photo_data = message.document

#     await message.answer(f'{photo_data}')

@dp.message(CommandStart())
async def cmd_start(message: Message):

    await message.answer(text=hello1_text.format(name=message.from_user.first_name), reply_markup=start_kb())

@dp.callback_query(F.data == "Intimissimi")
async def intimissimi(callback: CallbackQuery, state: FSMContext):
    wb = openpyxl.load_workbook("intimissimi.xlsx")
    user = [i[0] for i in wb['Sheet1'].values]
    if callback.from_user.id in user:
        await callback.message.answer(text="Вы уже зарегистрированы на это мероприятие")
    else:
        await state.set_state(Register.fio)
        await state.update_data(username=callback.message.from_user.username, tg_id=callback.message.from_user.id, types="intimissimi")
        await callback.message.answer(text="Расскажите немного о себе. Введите своё ФИО")

@dp.callback_query(F.data == "Интершарм")
async def intersharm(callback: CallbackQuery, state: FSMContext):
    wb = openpyxl.load_workbook("intersharm.xlsx")
    user = [i[0] for i in wb['Sheet1'].values]
    if callback.from_user.id in user:
        await callback.message.answer(text="Вы уже зарегестрированы на это мероприятие")
    else:
        await state.set_state(Register.fio)
        await state.update_data(username=callback.message.from_user.username, tg_id=callback.message.from_user.id, types="intersharm")
        await callback.message.answer(text="Расскажите немного о себе. Введите своё ФИО")

@dp.message(Register.fio)
async def reg_fio(message: Message, state: FSMContext) -> None:
    await state.update_data(fio=message.text)
    await state.set_state(Register.age)
    await message.answer(text="Введите свой возраст")

@dp.message(Register.age)
async def reg_age(message: Message, state: FSMContext) -> None:
    await state.update_data(age=message.text)
    await state.set_state(Register.city)
    await message.answer(text="В каком городе вы живете?")

@dp.message(Register.city)
async def reg_city(message: Message, state: FSMContext) -> None:
    await state.update_data(city=message.text)
    await state.set_state(Register.mail)
    await message.answer(text="Введите свою электронную почту")

@dp.message(Register.mail)
async def reg_mail(message: Message, state: FSMContext) -> None:
    if match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', message.text) is None:
        await message.answer(text="Неправильный формат данных, попробуйте снова")
        return
    
    await state.update_data(email=message.text)

    await state.set_state(Register.number)
    await message.answer(text="Отправьте свой номер", reply_markup=contact_keyboard())


@dp.message(Register.number)
async def reg_agree(message: Message, state: FSMContext) -> None:
    try:
        await state.update_data(number=message.contact.phone_number)
    except:
        await state.update_data(number=message.text)

    await message.answer_document(document=agree_doc, caption=agree_text, reply_markup=agree_kb(), parse_mode="HTML")
    
@dp.callback_query(F.data == "Отправить")
async def reg_final(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    tg_id = callback.from_user.id
    username = callback.from_user.username
    fio = data["fio"]
    age = data["age"]
    city = data["city"]
    email = data["email"]
    number = data["number"]
    types = data["types"]
    await state.clear()
    await create_user(tg_id, username, fio, age, city, email, number, types)
    await callback.message.answer(text=thx_text, reply_markup=ReplyKeyboardRemove())
    await callback.message.answer(text=main_text, reply_markup=main_menu_kb())

async def create_user(tg_id, username, fio, age, city, email, number, types):
    if types == "intimissimi":
        wb = openpyxl.load_workbook("intimissimi.xlsx")
    else:
        wb = openpyxl.load_workbook("intersharm.xlsx")
    sheet = wb["Sheet1"]
    sheet.append([tg_id, username, fio, age, city, email, number])
    wb.save(types + ".xlsx")

@dp.callback_query(F.data == "Главное меню")
async def main_menu(callback: CallbackQuery):
    await callback.message.delete()
    await callback.message.answer(text=main_text, reply_markup=main_menu_kb())

@dp.callback_query(F.data == "Линия")
async def hot_line(callback: CallbackQuery):
    await callback.message.answer(text=line_text, reply_markup=bot_kb())

@dp.callback_query(F.data == "Узнать")
async def more(callback: CallbackQuery):
    await callback.message.answer(text="Хорошо, вот, что я могу предложить", reply_markup=more_kb())

@dp.callback_query(F.data == "Спросить")
async def ask_doctor(callback: CallbackQuery):
    await callback.message.answer(text="Хоитите получить консультацию онкомаммолога?", reply_markup=consultation_kb())

@dp.callback_query(F.data == "Онлайн")
async def online(callback: CallbackQuery):
    await callback.message.answer(text='Онлайн марафон "5 шагов к здоровью груди"')
    
@dp.callback_query(F.data == "Очно")
async def ochno(callback: CallbackQuery):
    await callback.message.answer(text='Выберите город', reply_markup=city_kb())
    
@dp.callback_query(F.data == "Москва")
async def moscow(callback: CallbackQuery):
    await callback.message.answer(text=moscow_text, reply_markup=moscow_kb())
    
@dp.callback_query(F.data == "Европейский")
async def europe(callback: CallbackQuery):
    await callback.message.answer(text=europe_text, reply_markup=reg_intimissimi_kb())
    
@dp.callback_query(F.data == "Авиапарк")
async def aviapark(callback: CallbackQuery):
    await callback.message.answer(text=aviapark_text, reply_markup=reg_intimissimi_kb())
    
@dp.callback_query(F.data == "Метрополис")
async def metropolis(callback: CallbackQuery):
    await callback.message.answer(text=metropolis_text, reply_markup=reg_intimissimi_kb())
    
@dp.callback_query(F.data == "Выставка")
async def intersharm(callback: CallbackQuery):
    await callback.message.answer(text=intersharm_text, reply_markup=reg_intersharm_kb())

@dp.callback_query(F.data == "Воронеж")
async def voronezh(callback: CallbackQuery):
    await callback.message.answer(text=voronezh_text, reply_markup=menu_kb())
    
@dp.callback_query(F.data == "Казань")
async def kazan(callback: CallbackQuery):
    await callback.message.answer(text=kazan_text, reply_markup=menu_kb())
    
@dp.callback_query(F.data == "Нижний Новгород")
async def nizhnii(callback: CallbackQuery):
    await callback.message.answer(text=nizhnii_text, reply_markup=menu_kb())
    
@dp.callback_query(F.data == "Сочи")
async def sochi(callback: CallbackQuery):
    await callback.message.answer(text=sochi_text, reply_markup=menu_kb())
    
@dp.callback_query(F.data == "Помощь")
async def help(callback: CallbackQuery):
    await callback.message.answer(text="Хорошо, вот, что я могу предложить", reply_markup=help_kb())
    
@dp.callback_query(F.data == "Квест")
async def quest(callback: CallbackQuery):
    await callback.message.answer(text=quest_text, reply_markup=quest_kb())
    
@dp.callback_query(F.data == "Задание 1")
async def first_quest(callback: CallbackQuery):
    await callback.message.answer(text=first_quest_text, reply_markup=first_kb())
    
@dp.callback_query(F.data == "Задание 2")
async def second_quest(callback: CallbackQuery):
    await callback.message.answer(text=second_quest_text, reply_markup=second_kb())
    
@dp.callback_query(F.data == "Задание 3")
async def third_quest(callback: CallbackQuery):
    await callback.message.answer(text=third_quest_text, reply_markup=third_kb())
    
@dp.callback_query(F.data == "Задание 4")
async def firth_quest(callback: CallbackQuery):
    await callback.message.answer(text=firth_quest_text, reply_markup=menu_kb())
    





